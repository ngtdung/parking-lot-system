# Import libraries and functions from other python scripts
import PySimpleGUI as sg
import pandas as pd
from mfrc522 import SimpleMFRC522
import plate_detection as pld
from scan_ID_to_info import show_info
import logr as lr
import test_cam as tc
import time
import cv2
from openpyxl import Workbook, load_workbook
from picamera2 import Picamera2, Preview
from PIL import Image

# Import data
df = pd.read_excel(r'RFID.xlsx')
sg.theme('LightBlue') # Set theme

# Create layout
layout = [
    [sg.Button('Scan ID'), sg.Input(key='ID_scanned', do_not_clear=True, size=(50, 4)), sg.Text(key='ID_status')],
    [sg.Button('Exit')],
]

window = sg.Window('Create daily ticket', layout) # Initialize Window
            
# Function to get plate from log file
def extract_plate_from_log(ID, wb, row_value):
    ws = wb['Sheet1']
    if row_value > 0: # If the row_value > 0, the system will recognize it is a logout event and display both login and logout plates
        print('out')
        entry_plate = ws.cell(row=row_value, column=4).value
        exit_plate = ws.cell(row=row_value, column=5).value
        Image.open(entry_plate).show()
        Image.open(exit_plate).show()
    else: # Else display the entry plate only
        print('in')
        entry_plate = ws.cell(row=ws.max_row, column=4).value
        Image.open(entry_plate).show()
    
# Function to recognize if the card is in the system
def validate_id_code(ID_code):
    if ID_code == '':
        return 'Please enter a valid ID'
    else:
        value = int(ID_code)
        id_list = df.values[:, 5].tolist()
        if value in id_list:
            return 'Monthly ticket'
        else:
            return 'Daily ticket'


while True:
    event, values = window.read() # Get event and values
    if event == sg.WIN_CLOSED or event == 'Exit': # Terminate the program if the window closed or exit button pressed
        break
    elif event == 'Scan ID':
        wb = load_workbook(filename="Log.xlsx") # Import data from Log file
        reader = SimpleMFRC522() # Initialize RFID Reader
        ID_reader,text = reader.read() # Get card's ID
        row_val = lr.check(ID_reader, wb) # Get the row of the ID
        window['ID_scanned'].update(ID_reader)
        error_message = validate_id_code(ID_reader) # Validate the Card's ID
        if error_message == 'Please enter a valid ID':
            window['ID_status'].update(error_message)
        elif error_message == 'Monthly ticket': # Display the info of the card if it's a monthly card
            window['ID_status'].update(error_message)
            filtered_df = df[df['Output_ID'].notna()].drop(0, axis=1)
            row = filtered_df[filtered_df['Output_ID'] == int(ID_reader)].iloc[0]
            info = row.values.tolist()
            show_info(info)
        elif error_message == 'Daily ticket': # Display the message "Daily ticket"
            window['ID_status'].update(error_message)
        img_path = tc.take_picture(ID_reader) # Take picture of the license plate once the card is scanned
        pld.plate_retrieval(img_path, row_val)  # Extract the plate then save the photo
        lr.log(wb, ID_reader, img_path, row_val) # Log the ID, entry time, exit time, and plate's image path into log file
        extract_plate_from_log(ID_reader, wb, row_val) # Display the plates upon exit
window.close()
