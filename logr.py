# Import libraries
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
import time

# Function check if the event is login or logout
def check(id, wb):
    ws = wb['Sheet1']
    r = 0
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
    rows = reversed(rows)
    for row in rows:
        if row[0].value == id and row[2].value is None: # Match the ID and the event that hasn't logged out
            r = row[0].row # Assign that row's value to r
            break
    return r # r = 0 for login, r!=0 for logout

# Function to log ID, entry/exit time and plate image to log file
def log(wb, id, img_path, r):
    ws = wb['Sheet1']
    i = 0
    if r == 0: # Login
        ws.cell(row=ws.max_row + 1, column=1).value = id
        ws.cell(row=ws.max_row, column=2).value = time.ctime(time.time())
        ws.cell(row=ws.max_row, column=4).value = img_path
    else: # Logout
        ws.cell(row=r, column=3).value = time.ctime(time.time())
        ws.cell(row=r, column=5).value = img_path
        i = 1

    wb.save("Log.xlsx")

    return i
