from openpyxl import Workbook
from openpyxl import load_workbook
import time

wb = load_workbook(filename="Log.xlsx")

ws = wb["Sheet1"]


def log(id):
    i = 0
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
    rows = reversed(rows)
    for row in rows:
        if row[0].value == id and row[2].value is None:
            ws.cell(row=cell.row, column=3).value = time.ctime(time.time())
            i = 1
            break
    if i == 0:
        ws.cell(row=ws.max_row + 1, column=1).value = id
        ws.cell(row=ws.max_row, column=2).value = time.ctime(time.time())
    wb.save("Log.xlsx")







logout("29P188276")
